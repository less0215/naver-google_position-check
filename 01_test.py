import streamlit as st
import time
import random
import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import hashlib
import hmac
import base64
import traceback
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor


st.set_page_config(
    page_title="법무법인 동주 SEO",
    layout='wide'
)

# Naver API 관련 함수 및 설정
BASE_URL = 'https://api.naver.com'
API_KEY = '010000000094450d1dd02d9f94675fb0c3b77ee5d03ef32f1f0b956eae9cb19851dcb59d5b'
SECRET_KEY = 'AQAAAACURQ0d0C2flGdfsMO3fuXQj9OGFEyr4CjF7kcsHnhtOg=='
CUSTOMER_ID = '1943381'

# WebDriver 초기화 함수 정의
def initialize_webdriver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-features=NetworkService")
    chrome_options.add_argument("--window-size=1920x1080")
    chrome_options.add_argument("--disable-features=VizDisplayCompositor")
    
    # 추가된 옵션
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--lang=ko-KR")
    chrome_options.add_argument("--accept-language=ko-KR,ko;q=0.9")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

    try:
        driver = webdriver.Chrome(options=chrome_options)
        return driver
    except Exception as e:
        st.error(f"WebDriver 초기화 중 오류 발생: {str(e)}")
        st.info("관리자에게 문의하세요.")
        return None

class Signature:
    @staticmethod
    def generate(timestamp, method, uri, secret_key):
        message = "{}.{}.{}".format(timestamp, method, uri)
        hash = hmac.new(bytes(secret_key, "utf-8"), bytes(message, "utf-8"), hashlib.sha256)
        
        hash.hexdigest()
        return base64.b64encode(hash.digest())

# 전역 변수로 WebDriver 풀 생성
driver_pool = concurrent.futures.ThreadPoolExecutor(max_workers=5)

def get_google_search_results(keyword, dongju_url_dict):
    driver = driver_pool.submit(initialize_webdriver).result()
    if driver is None:
        return None, None

    results = {
        '키워드': keyword,
        '스니펫': '',
    }

    for i in range(1, 16):
        results[f'{i}'] = ''

    try:
        driver.get(f"https://www.google.com/search?q={keyword}")

        # 스니펫 확인
        try:
            snippet = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".g.wF4fFd.JnwWd.g-blk .tjvcx.GvPZzd.cHaqb"))
            )
            snippet_text = snippet.text.split('›')[0].strip()
            for url, name in dongju_url_dict.items():
                if url in snippet_text:
                    results['스니펫'] = name
                    break
        except:
            pass

        # 순위 확인
        links = WebDriverWait(driver, 5).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.g a'))
        )
        for i, link in enumerate(links[:15], start=1):
            href = link.get_attribute('href')
            for url, name in dongju_url_dict.items():
                if url in href:
                    results[f'{i}'] = name
                    break

        # 연관 검색어 추출
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        rel_keywords = soup.select(".oatEtb .dg6jd")
        related_keywords = [rel_keyword.text for rel_keyword in rel_keywords]

    except Exception as e:
        error_msg = traceback.format_exc()
        st.error(f"검색 중 오류 발생: {str(e)}")
        st.text(error_msg)
        st.info("오류가 지속되면 관리자에게 문의하세요.")
        related_keywords = []
    finally:
        driver_pool.submit(driver.quit)

    return results, related_keywords

def click_blog_tab(driver):
    selectors = [
        'a[role="tab"].tab:has(i.ico_nav_blog)[onclick*="a=tab*b.jmp"]',  # 새로운 선택자
        '.flick_bx:nth-of-type(1) > a',
        '.flick_bx:nth-of-type(2) > a',
        '.flick_bx:nth-of-type(3) > a',
        '[data-tab="view"][data-type="section"]'
    ]
    
    for selector in selectors:
        try:
            element = WebDriverWait(driver, 1.5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
            )
            element.click()
            time.sleep(1)  # 클릭 후 잠시 대기
            
            # 블로그 탭이 클릭되었는지 확인
            if "blog" in driver.current_url:
                return True
        except Exception as e:
            print(f"선택자 {selector}에 대한 클릭 실패: {str(e)}")
            continue
    
    print("블로그 탭을 찾을 수 없습니다.")
    return False

def process_smartblock_results(driver, dongju_id_list):
    extracted_ids = []
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    keywords = soup.select('.kmB6JnsyOzYVwnAzyoAL.fds-info-inner-text')

    for keyword in keywords:
        href = keyword.get('href', '').split('/')[3]
        
        if '?art' in href:
            extracted_id = href.split('?art')[0]
        else:
            extracted_id = href
        
        extracted_ids.append(extracted_id)

    matching_id = next((id for id in extracted_ids if id in dongju_id_list), None)
    
    return matching_id

def process_keywords(keyword_list, dongju_url_dict):
    results_list = []
    related_keywords_dict = {}
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_keyword = {executor.submit(get_google_search_results, keyword, dongju_url_dict): keyword for keyword in keyword_list}
        for future in concurrent.futures.as_completed(future_to_keyword):
            keyword = future_to_keyword[future]
            try:
                results, related_keywords = future.result()
                if results is not None:
                    results_list.append(results)
                    related_keywords_dict[keyword] = related_keywords
            except Exception as exc:
                st.error(f'{keyword} generated an exception: {exc}')
    
    return results_list, related_keywords_dict

# 스니펫 배경색 적용 함수
def highlight_snippet(val):
    if val:
        return 'background-color: #90EE90'
    return ''


# 구글 검색용 엑셀 파일 생성 함수
def create_excel_google(df):
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    # 헤더 추가
    for col, value in enumerate(df.columns.values, start=1):
        sheet.cell(row=1, column=col, value=value)

    # 데이터 추가 및 스타일 적용
    for row, (index, data) in enumerate(df.iterrows(), start=2):
        for col, value in enumerate(data.values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if col == 2:  # 스니펫 열
                if value:  # 값이 있는 경우에만 배경색 적용
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # 열 너비 자동 조정
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(output)
    return output.getvalue()

# 엑셀 파일 생성 함수
def create_excel(df):
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    # 헤더 추가
    for col, value in enumerate(df.columns.values, start=1):
        sheet.cell(row=1, column=col, value=value)

    # 데이터 추가 및 스타일 적용
    for row, (index, data) in enumerate(df.iterrows(), start=2):
        for col, value in enumerate(data.values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if col == 2:  # 스니펫 열
                if value:  # 값이 있는 경우에만 배경색 적용
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # 열 너비 자동 조정
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(output)
    return output.getvalue()


def get_header(method, uri, api_key, secret_key, customer_id):
    timestamp = str(round(time.time() * 1000))
    signature = Signature.generate(timestamp, method, uri, secret_key)
    
    return {
        'Content-Type': 'application/json; charset=UTF-8',
        'X-Timestamp': timestamp,
        'X-API-KEY': api_key,
        'X-Customer': str(customer_id),
        'X-Signature': signature
    }

def get_search_volume(keyword):
    uri = '/keywordstool'
    method = 'GET'
    params = {'hintKeywords': keyword, 'showDetail': '1'}
    
    try:
        r = requests.get(BASE_URL + uri, params=params, 
                         headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        r.raise_for_status()  # Raise an exception for bad status codes
        
        data = r.json()['keywordList']
        result = next((item for item in data if item['relKeyword'] == keyword), None)
        
        if result:
            return result['monthlyPcQcCnt'], result['monthlyMobileQcCnt']
        else:
            st.warning(f"검색량 데이터를 찾을 수 없습니다: {keyword}")
            return 0, 0
    except requests.exceptions.RequestException as e:
        st.error(f"검색량 조회 중 오류 발생: {str(e)}")
        return 0, 0
    except Exception as e:
        st.error(f"예상치 못한 오류 발생: {str(e)}")
        return 0, 0

def get_naver_search_results(keyword, dongju_id_list):
    keyword_type = ''

    driver = driver_pool.submit(initialize_webdriver).result()
    if driver is None:
        return None, None

    results = {
        '키워드': keyword,
        '스니펫': '',
        '스블': '',
        'M': 0,
        'P': 0
    }
    for i in range(1, 16):
        results[f'{i}'] = ''

    try:
        preprocessed_keyword = keyword.replace(" ", "")
        driver.get(f"https://search.naver.com/search.naver?ssc=tab.nx.all&where=nexearch&sm=tab_jum&query={preprocessed_keyword}")

        is_knowledge_snippet = False
        is_smartblock = False
        
        # 지식스니펫 확인
        try:
            knowledge_snippet = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.source_box .txt.elss'))
            )
            split_knowledge_snippet = knowledge_snippet.get_attribute('href').split('/')[3]
            is_knowledge_snippet = True
            if split_knowledge_snippet in dongju_id_list:
                results['스니펫'] = split_knowledge_snippet
        except:
            pass
        
        # 스마트블럭 확인 및 처리
        try:
            WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.gSQMmoVs7gF12hlu3vMg.desktop_mode.api_subject_bx'))
            )
            is_smartblock = True
            results['스블'] = process_smartblock_results(driver, dongju_id_list)
        except:
            pass
        
        # 키워드 유형 결정
        if is_knowledge_snippet and is_smartblock:
            keyword_type = 'both'
        elif is_knowledge_snippet:
            keyword_type = 'knowledge_snippet'
        elif is_smartblock:
            keyword_type = 'smartblock'
        else:
            keyword_type = 'normal'

        # 블로그 탭 클릭
        WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.flick_bx:nth-of-type(3) > a'))
        ).click()

        # 블로그 순위 체크
        blog_ids = WebDriverWait(driver, 1.5).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.user_info a'))
        )
        for rank, blog_id in enumerate(blog_ids[:15], start=1):
            href = blog_id.get_attribute('href')
            extracted_id = href.split('/')[-1]
            if extracted_id in dongju_id_list:
                results[f'{rank}'] = extracted_id

        # 검색량 조회 (모든 키워드 타입에 대해 수행)
        pc_volume, mobile_volume = get_search_volume(preprocessed_keyword)
        results['M'] = mobile_volume
        results['P'] = pc_volume

    except Exception as e:
        error_msg = traceback.format_exc()
        st.error(f"키워드 '{keyword}' 처리 중 오류 발생: {str(e)}")
        st.text(error_msg)
        st.info("오류가 지속되면 관리자에게 문의하세요.")
    finally:
        driver_pool.submit(driver.quit)

    return results, keyword_type

def process_keywords(keyword_list, dongju_id_list):
    results_list = []
    keyword_types = {}
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_keyword = {executor.submit(get_naver_search_results, keyword, dongju_id_list): keyword for keyword in keyword_list}
        for future in concurrent.futures.as_completed(future_to_keyword):
            keyword = future_to_keyword[future]
            try:
                results, keyword_type = future.result()
                if results is not None:
                    results_list.append(results)
                    keyword_types[keyword] = keyword_type
            except Exception as exc:
                st.error(f'{keyword} generated an exception: {exc}')
    
    return results_list, keyword_types

# 색상 적용 함수
def color_keyword(val, keyword_types, keyword, column_name):
    keyword_type = keyword_types.get(keyword, '')
    if column_name == '키워드':
        if keyword_type == 'knowledge_snippet':
            return 'background-color: #90EE90'  # 초록색
        elif keyword_type == 'smartblock':
            return 'background-color: #ADD8E6'  # 파란색
        elif keyword_type == 'both':
            return 'background-color: #FFB3BA'  # 빨간색
    elif column_name == '스니펫':
        if val:  # 값이 있을 때만 배경색 적용
            return 'background-color: #90EE90'  # 초록색
    elif column_name == '스블':
        if val:  # 값이 있을 때만 배경색 적용
            return 'background-color: #ADD8E6'  # 파란색
    return ''

# 엑셀 파일 생성 함수 수정
def create_excel(df, keyword_types, smartblock_keywords):
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    # 헤더 추가
    for col, value in enumerate(df.columns.values, start=1):
        sheet.cell(row=1, column=col, value=value)

    # 데이터 추가 및 스타일 적용
    for row, (index, data) in enumerate(df.iterrows(), start=2):
        keyword = data['키워드']
        keyword_type = keyword_types.get(keyword, '')
        for col, value in enumerate(data.values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if col == 1:  # 키워드 열
                if keyword_type == 'knowledge_snippet':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif keyword_type == 'smartblock':
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif keyword_type == 'both':
                    cell.fill = PatternFill(start_color="FFB3BA", end_color="FFB3BA", fill_type="solid")
            elif col == 2:  # 스니펫 열
                if value:  # 값이 있는 경우에만 배경색 적용
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif col == 3:  # 스블 열
                if value:  # 값이 있는 경우에만 배경색 적용
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # 스마트블럭 키워드 및 연관 키워드 추가
    sheet = workbook.create_sheet(title="스마트블럭 키워드")
    sheet.append(["스마트블럭 키워드", "연관 키워드"])
    for keyword, related_keywords in smartblock_keywords.items():
        sheet.append([keyword, ", ".join(related_keywords)])

    # 열 너비 자동 조정
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(output)
    return output.getvalue()

# 키워드 전처리 함수 추가
def preprocess_keyword(keyword):
    return keyword.replace(" ", "")

# 사이드탭 생성
selected_tab = st.sidebar.radio("검색 엔진 선택", ["네이버", "구글"])

if selected_tab == "네이버":
    # 네이버 탭 내용
    st.title("🔍 네이버 순위 체크 및 검색량 조회")

    # 팀 선택
    selected_team = st.selectbox("팀 선택", ["청소년팀", "형사팀", "경제팀", "신규팀(음주&고소대리)"])

    # 키워드 입력
    keywords = st.text_area("키워드를 입력해 주세요 (한 줄에 하나씩)", height=200)

    # 동주 ID 리스트 (업데이트됨)
    dongju_id_list = [
        # 청소년팀
        "designersiun", "singsong0514", "phoenixjeong", "hamas3000", "roses777",
        "dongjulaw1", "dongjulaw2", "dongjusuwon1", "dongjulaw6", "dj_ehdwn1",
        # 형사팀
        "rudnfdldi00", "ehtlarhdwn", "widance", "yellowoi", "dongjulaw",
        "tale1396", "dongjulaw5", "dongjulaw100", "dongjulaw4", "dongjulaw02",
        # 경제팀
        "dksro018",
        # 신규팀(음주&고소대리)
        "cckjjt", "qusghtkehdwn", "dongjulaw7", "ujm159",
        # 기타 ID (기존에 있던 ID들)
        "dong-ju-law", "dongjulaw3", "ehdwnfh", "kkobugi39"
    ]

    # 순위 확인 버튼
    if st.button("순위 확인"):
        if not keywords:
            st.error("키워드를 입력해주세요.")
        else:
            keyword_list = [keyword.strip() for keyword in keywords.split('\n') if keyword.strip()]

            if not keyword_list:
                st.error("유효한 키워드를 입력해주세요.")
            else:
                driver = initialize_webdriver()
                if driver is None:
                    st.stop()

                try:
                    results_list = []
                    keyword_types = {}
                    smartblock_keywords = {}

                    result_placeholder = st.empty()
                    progress_bar = st.progress(0)

                    # 스타일 정의 부분 (기존 코드 유지)
                    st.markdown("""
                    <style>
                        .color-box {
                            padding: 10px;
                            border-radius: 4px;
                            margin-bottom: 10px;
                        }
                        .color-box p {
                            margin: 0;
                            font-size: 16px;
                            text-align: center;
                        }
                        .section-header {
                            font-size: 20px;
                            font-weight: bold;
                            margin-bottom: 15px;
                        }
                    </style>
                    """, unsafe_allow_html=True)

                    # 각 키워드에 대해 검색 수행
                    for i, keyword in enumerate(keyword_list):
                        try:
                            preprocessed_keyword = preprocess_keyword(keyword)
                            driver.get(f"https://search.naver.com/search.naver?ssc=tab.nx.all&where=nexearch&sm=tab_jum&query={preprocessed_keyword}")

                            keyword_type = ''
                            is_knowledge_snippet = False
                            is_smartblock = False

                            # 지식스니펫 확인
                            snippet_id = ''
                            try:
                                knowledge_snippet = driver.find_element(By.CSS_SELECTOR, '.source_box .txt.elss').get_attribute('href')
                                split_knowledge_snippet = knowledge_snippet.split('/')[3]
                                is_knowledge_snippet = True
                                if split_knowledge_snippet in dongju_id_list:
                                    snippet_id = split_knowledge_snippet
                            except:
                                pass
                            
                            # 스마트블럭 확인 및 처리
                            smartblock_id = ''
                            try:
                                smartblock_research = driver.find_element(By.CSS_SELECTOR, '.gSQMmoVs7gF12hlu3vMg.desktop_mode.api_subject_bx')
                                is_smartblock = True
                                smartblock_id = process_smartblock_results(driver, dongju_id_list)
                            except:
                                pass
                            
                            # 키워드 유형 결정
                            if is_knowledge_snippet and is_smartblock:
                                keyword_type = 'both'
                            elif is_knowledge_snippet:
                                keyword_type = 'knowledge_snippet'
                            elif is_smartblock:
                                keyword_type = 'smartblock'

                            # 키워드 유형 저장
                            keyword_types[keyword] = keyword_type

                            # 여기서부터 새로운 코드를 삽입합니다
                            try:
                                if not click_blog_tab(driver):
                                    st.warning(f"키워드 '{keyword}'에 대한 블로그 탭을 찾을 수 없습니다.")
                                    continue
                                
                                # 3번 스크롤 처리
                                for _ in range(3):
                                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                    time.sleep(random.uniform(1, 1.5))

                            except Exception as e:
                                st.error(f"키워드 '{keyword}' 처리 중 오류 발생: {str(e)}")
                                continue
                            
                            # 블로그 순위 체크 (기존 코드)
                            blog_ids = driver.find_elements(By.CSS_SELECTOR, '.user_info a')
                            results = {j: '' for j in range(1, 16)}  # 모든 순위를 빈 문자열로 초기화
                            for rank, blog_id in enumerate(blog_ids, start=1):
                                if rank > 15:  # 15위까지만 체크
                                    break
                                href = blog_id.get_attribute('href')
                                extracted_id = href.split('/')[-1]
                                if extracted_id in dongju_id_list:
                                    results[rank] = extracted_id

                            # 검색량 조회
                            pc_volume, mobile_volume = get_search_volume(preprocessed_keyword)

                            # 결과 리스트에 추가
                            row = {'키워드': keyword, '스니펫': snippet_id, '스블': smartblock_id, 'M': mobile_volume, 'P': pc_volume}
                            row.update(results)
                            results_list.append(row)

                            # 실시간으로 결과 표시
                            with result_placeholder.container():
                                st.markdown('<p class="section-header">실시간 검색 결과</p>', unsafe_allow_html=True)
                                df = pd.DataFrame(results_list)
                                styled_df = df.style.apply(lambda row: [color_keyword(val, keyword_types, row['키워드'], col) for col, val in row.items()], axis=1)
                                st.dataframe(styled_df, use_container_width=True)

                                st.markdown("<br>", unsafe_allow_html=True)

                                st.markdown('<p class="section-header">키워드 배경색 설명</p>', unsafe_allow_html=True)
                                col1, col2, col3, col4 = st.columns(4)

                                with col1:
                                    st.markdown(
                                        """
                                        <div class="color-box" style="background-color: #FFB3BA;">
                                            <p style="text-align: left;">지식스니펫 + 스마트블럭</p>
                                        </div>
                                        """,
                                        unsafe_allow_html=True
                                    )

                                with col2:
                                    st.markdown(
                                        """
                                        <div class="color-box" style="background-color: #90EE90;">
                                            <p style="text-align: left;">지식스니펫</p>
                                        </div>
                                        """,
                                        unsafe_allow_html=True
                                    )

                                with col3:
                                    st.markdown(
                                        """
                                        <div class="color-box" style="background-color: #ADD8E6;">
                                            <p style="text-align: left;">스마트블럭</p>
                                        </div>
                                        """,
                                        unsafe_allow_html=True
                                    )

                                with col4:
                                    st.markdown(
                                        """
                                        <div class="color-box" style="background-color: #F0F2F6;">
                                            <p style="text-align: left;">일반키워드</p>
                                        </div>
                                        """,
                                        unsafe_allow_html=True
                                    )

                                if smartblock_keywords:
                                    st.markdown("<br>", unsafe_allow_html=True)
                                    st.markdown('<p class="section-header">스마트블럭 키워드 및 연관 키워드</p>', unsafe_allow_html=True)
                                    for kw, related_kws in smartblock_keywords.items():
                                        with st.expander(f"키워드: {kw}"):
                                            st.write(f"연관 키워드: {', '.join(related_kws)}")

                            # 진행 상황 업데이트
                            progress_bar.progress((i + 1) / len(keyword_list))

                            # 각 키워드 검색 후 잠시 대기
                            time.sleep(random.uniform(1, 1.5))

                        except Exception as e:
                            error_msg = traceback.format_exc()
                            st.error(f"키워드 '{keyword}' 처리 중 오류 발생: {str(e)}")
                            st.text(error_msg)
                            st.info("오류가 지속되면 관리자에게 문의하세요.")

                finally:
                    if driver:
                        driver.quit()

                # 엑셀 다운로드 버튼
                excel_data = create_excel(df, keyword_types, smartblock_keywords)
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=excel_data,
                    file_name="search_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    st.info("'순위 확인' 버튼을 클릭해서 검색 결과를 실시간으로 확인하세요.")

def process_keywords(keyword_list, dongju_url_dict):
    results_list = []
    related_keywords_dict = {}
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_keyword = {executor.submit(get_google_search_results, keyword, dongju_url_dict): keyword for keyword in keyword_list}
        for future in concurrent.futures.as_completed(future_to_keyword):
            keyword = future_to_keyword[future]
            try:
                results, related_keywords = future.result()
                if results is not None:
                    results_list.append(results)
                    related_keywords_dict[keyword] = related_keywords
            except Exception as exc:
                st.error(f'{keyword} generated an exception: {exc}')
    
    return results_list, related_keywords_dict


# 구글 탭 내의 코드를 다음과 같이 수정
if selected_tab == "구글":
    st.title("🔍 구글 순위 체크")

    st.markdown("""
    클라우드 문제로 인해 순위 체크가 제대로 되지 않습니다.

    이에 따라 기존에 사용하던 기능 배포를 중단했습니다.

    구글 순위 체크 프로그램 사용을 희망하시는 분은 [이 링크](https://drive.google.com/file/d/1QRWs6MGMDHbpf_0XShA3nuL_t79ALzDY/view?usp=sharing)를 클릭해 주세요.
    """, unsafe_allow_html=True)
